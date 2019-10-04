#!/usr/bin/env python3
# -*-encoding: utf-8-*-
# author: Valentyn Kofanov


from tkinter import *


import openpyxl as xl
from openpyxl.chart import BarChart, Reference, Series


class IterRowDict:
    def __init__(self, wb, ws_name=None):
        if not ws_name:
            self.ws = wb.active
        else:
            self.ws = wb[ws_name]
        self.labels = None
        self.row_iterator = None

    def __iter__(self):
        self.row_iterator = self.ws.rows
        self.labels = list(map(lambda x: x.value, next(self.row_iterator)))
        return self

    def __next__(self):
        return {key: value for key, value in zip(self.labels, list(map(lambda x: x.value, next(self.row_iterator))))}


def get_names(filename):
    ws = xl.load_workbook(filename)["Products"]
    return list(map(lambda x:x.value, list(ws.columns)[1][1:ws.max_row]))

def get_date(filename):
    ws = xl.load_workbook(filename)["Sales"]
    return list(map(lambda x: str(x.value)[6:], list(ws.columns)[1][1:ws.max_row]))


def produce(filename, name, date):
    wb = xl.load_workbook(filename)
    rez = False
    for sheet in wb:
        if sheet.title == name + date:
            ws = wb[name + date]
            rez = True
            break
    if not rez:
        ws = wb.create_sheet(title=name + date)
    for i in range(1, 13):
        ws.cell(row=1, column=i).value = i
        ws.cell(row=2, column=i).value = 0

    for prod in IterRowDict(wb, "Products"):
        for sale in IterRowDict(wb, "Sales"):
            if prod["Name"] == name and sale["Date"][6:] == date and prod["id"] == prod["id"]:
                amount = int(prod["Price"]) * int(sale["Quantity"])
                ws.cell(row=2, column=int(sale["Date"][3:5])).value = amount

    ws = wb[name+date]
    chart = BarChart()
    x = Reference(ws, min_col=1, max_col=12, min_row=1)
    y = Reference(ws, min_col=1, max_col=12, min_row=2)
    chart.append(Series(y))
    chart.set_categories(x)
    ws.add_chart(chart, "A10")
    wb.save(filename)

def main_tk(filename, names_list, years_list):


    def get_data():
        try:
            n = str(lbs_names.get(lbs_names.curselection()))
            y = str(lbs_years.get(lbs_years.curselection()))
            produce(filename, n, y)
            top.quit()

        except :
            pass




    top = Tk()

    lbs_names = Listbox(top, exportselection=0)

    for num, item in enumerate(names_list):
        lbs_names.insert(num, item)


    lbs_years = Listbox(top, exportselection=0)

    for num, item in enumerate(years_list):
        lbs_years.insert(num, item)


    b = Button(top, text="OK", command=get_data)

    lbs_names.grid(row=0, column=0)
    lbs_years.grid(row=0, column=2)
    b.grid(row=1, column=1)



    top.mainloop()



def main(filename="orkbook1.xlsx"):


    main_tk(filename, list(set(get_names(filename))), list(set(get_date(filename))))




if __name__ == '__main__':
    main()
